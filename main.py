import tkinter as tk  # python 3
import datetime
import gspread
import pandas as pd
import openpyxl
import time
import socket


class SampleApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        # App Window Size::
        self.attributes('-fullscreen', True)

        # Create StatusBar
        self.status_bar = StatusBar(self)
        self.status_bar.pack(side="top", fill="x")
        # Create and place the connection status label
        # self.connection_status_label = tk.Label(self.status_bar, font=("Helvetica", 15))
        # self.connection_status_label.pack(side="right", padx=15, pady=(30, 0))

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (StartPage, UserStatusPage, MainUserPage, AdminPage, NotificationPage,
                  BorrowBookPage, ReturnBookPage, LoadingPage, StartingUpPage, IDScanLoadingPage,
                  BorrowBookLoadingPage, ReturnBookLoadingPage, TransactionsLoadingPage,
                  AutomaticLogOutPage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        frame = self.frames['StartingUpPage']
        frame.tkraise()
        self.update()

        # connection to database:
        self.db_url = ""
        self.gc = None
        self.db = None
        self.no_wifi_connection = True
        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            self.no_wifi_connection = False
            # initialize communication with the cloud database:
            self.db_url = "https://docs.google.com/spreadsheets/d/144bmhnqKytJMZwtBWR0IJ_UFbGy4gWWqukEfHV6laEU/edit?usp=sharing"
            self.gc = gspread.service_account(
                filename="/home/amermasarweh/Desktop/project/service_account.json")
            self.db = self.gc.open_by_url(self.db_url)
            # Synchorinize with local database:
            TransactionsWorkSheet = self.db.worksheet("Transactions")
            TransactionsWorkSheet.clear()
            workbook = openpyxl.load_workbook('/home/amermasarweh/Desktop/project/local_db.xlsx')

            # Select the worksheet (replace 'Sheet1' with the actual sheet name if different)
            transactions_worksheet = workbook["Transactions"]
            for row in transactions_worksheet.iter_rows(min_row=1, max_col=4):
                row_cells = [cell.value for cell in row]
                TransactionsWorkSheet.append_row(row_cells)
        # else: no_wifi_connection: status already states that

        # Variable to store Job Id of the currently scheduled logout job
        self.auto_logout_job = ""
        # Variable that stores the current state of the system to enable/disable auto logouts
        self.logged_in = False

        # Scheduled System Backup ( Copy google sheets to local Excel file):
        self.after(ms=600 * 1000, func=self.backup_data)  # backup every 10 minutes

        # db:
        self.Users = []

        self.Books = []

        self.Transactions = []

        self.frames["StartPage"].username_entry.focus_set()
        self.show_frame("StartPage")

    def update_connection_status(self, connected):
        if connected:
            self.status_bar.connection_status_label.config(text="Connected to WI-FI", fg="green")
        else:
            self.status_bar.connection_status_label.config(text="❌ No Connection", fg="red")
        self.update()

    def check_wifi(self):
        try:
            socket.create_connection(("8.8.8.8", 53), timeout=1)
            self.update_connection_status(connected=True)
            return True  # Connected to internet, likely Wi-Fi
        except OSError:
            self.update_connection_status(connected=False)
            return False

    def convert_string(self, s):
        if not s[0].isnumeric():
            substring = s[1:10]
            result = substring
            return result
        else:
            return s

    def backup_data(self):
        # Schedule the next backup procedure:
        self.after(ms=600 * 1000, func=self.backup_data)  # every 10 minutes
        connected_to_wifi_currently = self.check_wifi()
        if self.logged_in:
            return  # don't want to bother the user
        self.show_frame('LoadingPage')

        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            # if there's a WI-FI connection, load data from Google Sheets to local Excel file:
            sheets = ["Users", "Books", "Transactions", "Admins"]
            google_sheets = [(sheet_name, self.db.worksheet(sheet_name).id) for sheet_name in sheets]
            excel_file = "/home/amermasarweh/Desktop/project/local_db.xlsx"  # Replace with path to your Excel file
            with pd.ExcelWriter(
                    path=excel_file,
                    mode="w",
                    engine="openpyxl",
            ) as writer:
                for google_sheet in google_sheets:
                    sheet_url = f"https://docs.google.com/spreadsheets/d/144bmhnqKytJMZwtBWR0IJ_UFbGy4gWWqukEfHV6laEU/" \
                                f"export?format=csv&gid={google_sheet[1]}"
                    df = pd.read_csv(sheet_url)
                    df.to_excel(writer, sheet_name=google_sheet[0], index=False)
        else:
            self.no_wifi_connection = True

        self.show_frame('StartPage')

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        # update current page for auto log out
        if page_name != 'AutomaticLogOutPage':
            self.frames['AutomaticLogOutPage'].prev_page = page_name
        # delay the automatic logout since the user is interacting with the app:
        if page_name != 'StartPage' and self.logged_in is True:
            self.after_cancel(id=self.auto_logout_job)
            self.auto_logout_job = self.after(ms=60 * 1000, func=self.show_logout_warning)
        frame = self.frames[page_name]
        frame.tkraise()
        self.update()

    def validate_login(self, id):
        if id == "":
            self.show_notification(notification="Empty Fields")
            self.show_frame('StartPage')
            return
        id = self.convert_string(id)
        self.show_frame('IDScanLoadingPage')

        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            # check if the Admin is logging in:
            Admins = self.db.worksheet("Admins").get_all_records()
            admin_info = None
            for a in Admins:
                if str(a['admin_id']) == self.remove_leading(id):
                    admin_info = a
                    break
            if admin_info is not None:
                self.logged_in = True
                self.auto_logout_job = self.after(ms=60 * 1000, func=self.show_logout_warning)
                self.goto_admin_page()
                return

            # if not the Admin, then check if one of the standard users:
            # get users list from cloud database:
            self.Users = self.db.worksheet("Users").get_all_records()
            user_info = None
            for u in self.Users:
                if str(u['user_id']) == self.remove_leading(id):
                    user_info = u
                    break
            if user_info is None:
                self.show_notification(notification="User does not exist!")
                self.frames['StartPage'].username_entry.delete(0, tk.END)
                self.show_frame('StartPage')
            else:
                self.logged_in = True
                self.auto_logout_job = self.after(ms=60 * 1000, func=self.show_logout_warning)
                self.frames['MainUserPage'].user_id = str(user_info['user_id'])
                self.show_frame("MainUserPage")

        else:  # In case there's no WI-FI connection:
            self.no_wifi_connection = True
            excel_file_path = "/home/amermasarweh/Desktop/project/local_db.xlsx"
            # Open the workbook
            workbook = openpyxl.load_workbook(excel_file_path)

            # first, check if the Admin is logging in:
            admins_worksheet = workbook['Admins']
            admin_info = None
            for a in admins_worksheet.iter_rows(min_row=2, max_col=1):
                cell_value = a[0].value
                if str(cell_value) == self.remove_leading(id):
                    self.no_wifi_connection = True
                    self.show_notification('No WI-FI Connection!' + '\n' +
                                           'Please add the book manually via Google Sheets')
                    self.frames['StartPage'].username_entry.delete(0, tk.END)
                    self.show_frame('StartPage')
                    return

            # if not the Admin, check if it's one of the standard users:
            users_worksheet = workbook["Users"]
            user_info = None
            for row in users_worksheet.iter_rows(min_row=2, max_col=1):  # Skip the header row (assuming row 1)
                cell_value = row[0].value  # Access by index (0-based)
                if str(cell_value) == self.remove_leading(id):
                    user_info = cell_value
                    break
            if user_info is None:
                self.show_notification(notification="User does not exist!")
                self.frames['StartPage'].username_entry.delete(0, tk.END)
                self.show_frame('StartPage')
            else:
                self.logged_in = True
                self.auto_logout_job = self.after(ms=60 * 1000, func=self.show_logout_warning)
                self.frames['MainUserPage'].user_id = str(user_info)
                self.show_frame("MainUserPage")

    def logout(self):
        ''' clear login info from previous users and go back to start page: '''
        # cancel the currently scheduled auto log out job:
        self.after_cancel(id=self.auto_logout_job)
        # update system login status:
        self.logged_in = False
        # Log user out and go back to start page:
        self.frames["StartPage"].username_entry.delete(0, tk.END)
        self.frames['StartPage'].username_entry.focus_set()
        self.show_frame("StartPage")
        self.check_wifi()

    def goto_user_status_page(self, user_id, prev_page):
        self.show_frame('TransactionsLoadingPage')
        ########################################################################################
        # Empty the existing list to add fresh data:
        book_name_listbox = self.frames["UserStatusPage"].user_book_names_listbox
        book_name_listbox.delete(0, 'end')

        date_listbox = self.frames["UserStatusPage"].user_date_listbox
        date_listbox.delete(0, 'end')

        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            self.Transactions = self.db.worksheet("Transactions").get_all_records()

            for t in self.Transactions:
                if str(t['user_id']) == self.remove_leading(user_id):
                    book_name_listbox.insert('end', str(t['book_name']))
                    date_listbox.insert('end', str(t['date']))
        else:
            self.no_wifi_connection = True
            excel_file_path = "local_db.xlsx"
            # Open the workbook
            workbook = openpyxl.load_workbook(excel_file_path)
            # Select the worksheet (replace 'Sheet1' with the actual sheet name if different)
            transactions_worksheet = workbook["Transactions"]
            for row in transactions_worksheet.iter_rows(min_row=2, max_col=4):
                cell_value = row[0].value
                if str(cell_value) == self.remove_leading(user_id):
                    # listbox.insert('end', "Book Name: " + str(row[2].value) + " Date: " + str(row[3].value))
                    pass

        self.frames["UserStatusPage"].back_page = prev_page
        self.show_frame("UserStatusPage")

    def goto_admin_page(self):
        self.frames["AdminPage"].barcode_entry.delete(0, tk.END)
        self.frames["AdminPage"].barcode_entry.focus_set()
        self.show_frame("AdminPage")

    def goto_borrow_book_page(self, user_id):
        self.frames["BorrowBookPage"].user_id = user_id
        self.frames["BorrowBookPage"].barcode_entry.delete(0, tk.END)
        self.frames["BorrowBookPage"].barcode_entry.focus_set()
        self.show_frame("BorrowBookPage")

    def goto_return_book_page(self):
        self.frames["ReturnBookPage"].barcode_entry.delete(0, tk.END)
        self.frames["ReturnBookPage"].barcode_entry.focus_set()
        self.show_frame("ReturnBookPage")

    def return_book(self, barcode):
        self.show_frame('ReturnBookLoadingPage')
        ########################################################################################
        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            self.Transactions = self.db.worksheet("Transactions").get_all_records()
            TransactionsTable = self.db.worksheet("Transactions")
            index_to_delete = -1
            for index, t in enumerate(self.Transactions):
                if str(t['barcode']) == self.remove_leading(barcode):
                    index_to_delete = index
                    break
            if index_to_delete == -1:
                self.show_notification(notification="This Book Copy Has Not Been Borrowed Before!")
                self.frames['ReturnBookPage'].barcode_entry.delete(0, tk.END)
                self.frames['ReturnBookPage'].barcode_entry.focus_set()
                self.show_frame('ReturnBookPage')
                return
            TransactionsTable.delete_rows(index_to_delete + 2, index_to_delete + 2)

        else:
            self.no_wifi_connection = True
            excel_file_path = "/home/amermasarweh/Desktop/project/local_db.xlsx"
            # Open the workbook
            workbook = openpyxl.load_workbook(excel_file_path)

            # Select the worksheet (replace 'Sheet1' with the actual sheet name if different)
            transactions_worksheet = workbook["Transactions"]
            index_to_delete = -1
            for index, row in enumerate(transactions_worksheet.iter_rows(min_row=2, max_col=2)):
                cell_value = row[1].value
                if str(cell_value) == self.remove_leading(barcode):
                    index_to_delete = index
                    break
            if index_to_delete == -1:
                self.show_notification(notification="This Book Copy Has Not Been Borrowed Before!")
                self.frames['ReturnBookPage'].barcode_entry.delete(0, tk.END)
                self.frames['ReturnBookPage'].barcode_entry.focus_set()
                self.show_frame('ReturnBookPage')
                return
            transactions_worksheet.delete_rows(idx=index_to_delete + 2, amount=1)
            workbook.save(filename=excel_file_path)

        self.show_notification(notification="Book Returned Successfully!")
        self.frames["ReturnBookPage"].barcode_entry.delete(0, tk.END)
        self.frames["ReturnBookPage"].barcode_entry.focus_set()
        self.show_frame('ReturnBookPage')

    def remove_leading(self, barcode):
        return barcode.lstrip('0')

    def borrow_book(self, barcode, user_id):
        self.show_frame('BorrowBookLoadingPage')
        #####################################################################################################
        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            # get data from Google Sheets:
            self.Books = self.db.worksheet("Books").get_all_records()
            self.Transactions = self.db.worksheet("Transactions").get_all_records()

            # Check if the copy is available for borrow:
            book_info = None
            for b in self.Books:
                if str(b['barcode']) == self.remove_leading(barcode):
                    book_info = b
                    break
            if book_info == None:
                self.show_notification(notification="Book does not belong to the Library!")
                self.frames['BorrowBookPage'].barcode_entry.delete(0, tk.END)
                self.frames['BorrowBookPage'].barcode_entry.focus_set()
                self.show_frame('BorrowBookPage')
                return
            for t in self.Transactions:
                if str(t['barcode']) == self.remove_leading(barcode):
                    self.show_notification(notification="This Book Copy has not been returned yet!")
                    self.frames['BorrowBookPage'].barcode_entry.delete(0, tk.END)
                    self.frames['BorrowBookPage'].barcode_entry.focus_set()
                    self.show_frame('BorrowBookPage')
                    return

            # Else, create and enter a new transaction record for user_id and book_barcode:
            transaction_date = datetime.date.today()
            transaction_date_str = transaction_date.strftime("%B %d, %Y")
            new_row_data = [int(user_id), int(barcode), book_info['book_name'], transaction_date_str]
            TransactionsWorkSheet = self.db.worksheet("Transactions")
            TransactionsWorkSheet.append_row(new_row_data)

        else:  # there is no WI-FI connection currently:
            self.no_wifi_connection = True
            excel_file_path = "/home/amermasarweh/Desktop/project/local_db.xlsx"
            # Open the workbook
            workbook = openpyxl.load_workbook(excel_file_path)

            # Select the worksheet (replace 'Sheet1' with the actual sheet name if different)
            books_worksheet = workbook["Books"]
            transactions_worksheet = workbook["Transactions"]

            book_info = None
            for row in books_worksheet.iter_rows(min_row=2, max_col=2):  # Skip the header row (assuming row 1)
                cell_value = row
                if str(cell_value[0].value) == self.remove_leading(barcode):
                    book_info = cell_value[1].value
                    break
            if book_info is None:
                self.show_notification(notification="Book does not belong to the Library!")
                self.frames['BorrowBookPage'].barcode_entry.delete(0, tk.END)
                self.frames['BorrowBookPage'].barcode_entry.focus_set()
                self.show_frame('BorrowBookPage')
                return
            for row in transactions_worksheet.iter_rows(min_row=2, max_col=1):
                cell_value = row[0].value
                if str(cell_value) == self.remove_leading(barcode):
                    self.show_notification(notification="This Book Copy has not been returned yet!")
                    self.frames['BorrowBookPage'].barcode_entry.delete(0, tk.END)
                    self.frames['BorrowBookPage'].barcode_entry.focus_set()
                    self.show_frame('BorrowBookPage')
                    return

            # Else, create and enter a new transaction record for user_id and book_barcode:
            transaction_date = datetime.date.today()
            transaction_date_str = transaction_date.strftime("%B %d, %Y")
            new_row_data = [int(user_id), int(barcode), book_info, transaction_date_str]
            transactions_worksheet.append(new_row_data)
            workbook.save(filename=excel_file_path)

        self.show_notification(notification="Book Borrowed Successfully :)")
        self.frames["BorrowBookPage"].barcode_entry.delete(0, tk.END)
        self.frames["BorrowBookPage"].barcode_entry.focus_set()
        self.show_frame('BorrowBookPage')

    def add_book(self, barcode):
        self.show_frame('LoadingPage')
        #####################################################################################################
        connected_to_wifi_currently = self.check_wifi()
        if connected_to_wifi_currently:
            # If control reached here, then there is a WI-FI connection:
            if self.no_wifi_connection:
                self.sync_excel_to_google_sheet()  # in case there was no WI-FI, sync back all data
                self.no_wifi_connection = False

            # Check if the copy already exists in the database:
            self.Books = self.db.worksheet("Books").get_all_records()
            for b in self.Books:
                if str(b['barcode']) == self.remove_leading(barcode):
                    self.show_notification(notification="Book is already in the Library!")
                    self.frames["AdminPage"].barcode_entry.delete(0, tk.END)
                    self.frames["AdminPage"].barcode_entry.focus_set()
                    self.show_frame('AdminPage')
                    return

            # Else, create a new row in the Books table and add barcode
            new_row_data = [int(barcode), 'N/A']
            BooksWorkSheet = self.db.worksheet("Books")
            BooksWorkSheet.append_row(new_row_data)
            self.show_notification(notification="Book Added Successfully :)")
            self.frames["AdminPage"].barcode_entry.delete(0, tk.END)
            self.frames["AdminPage"].barcode_entry.focus_set()
            self.show_frame('AdminPage')

        else:
            # control reached here so there is no WI-FI connection:
            self.no_wifi_connection = True
            self.show_notification('No WI-FI Connection!')
            self.logout()

    def show_logout_warning(self):
        # Schedule the next automatic logout job:
        self.auto_logout_job = self.after(ms=60 * 1000, func=self.show_logout_warning)

        # Set timer to logout in 15 seconds if we got no user response:
        self.frames['AutomaticLogOutPage'].sched_logout = self.after(ms=15000, func=self.countdown_logout)

        # Now show the automatic logout message to user:
        self.show_frame('AutomaticLogOutPage')

    def countdown_logout(self):
        self.show_notification(notification="User Logged Out Automatically")
        self.logout()

    def show_notification(self, notification):
        self.frames['NotificationPage'].title_label.config(text='\n' + '\n' + '\n' + notification,
                                                           font=('Helvetica', 25, 'bold'))
        self.show_frame('NotificationPage')
        time.sleep(2)

    def sync_excel_to_google_sheet(self):

        # if control reaches here: then there is WI-FI currently
        # check if there was no WI-FI connection at system start-up:
        if self.db is None:
            self.db_url = "https://docs.google.com/spreadsheets/d/144bmhnqKytJMZwtBWR0IJ_UFbGy4gWWqukEfHV6laEU/edit?usp=sharing"
            self.gc = gspread.service_account(
                filename="/home/amermasarweh/Desktop/project/service_account.json")
            self.db = self.gc.open_by_url(self.db_url)

        # if control reaches here: this means there was no WI-FI connection and now it came back
        TransactionsWorkSheet = self.db.worksheet("Transactions")
        excel_file_path = '/home/amermasarweh/Desktop/project/local_db.xlsx'
        self.no_wifi_connection = False  # update connection status
        TransactionsWorkSheet.clear()
        workbook = openpyxl.load_workbook(excel_file_path)

        # Select the worksheet
        transactions_worksheet = workbook["Transactions"]
        for row in transactions_worksheet.iter_rows(min_row=1, max_col=4):
            row_cells = [cell.value for cell in row]
            TransactionsWorkSheet.append_row(row_cells)


class StartPage(tk.Frame):
    ''' Login Page: '''

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        # Label with larger font and padding
        label = tk.Label(self, text="Please Scan your ID Card or Enter Manually 🙂", font=("Arial", 20, "bold"))
        label.pack(side="top", fill="x", pady=20)  # Increased padding

        # Username label with smaller font

        # Entry with increased width
        self.username_entry = tk.Entry(self, width=30, font=("Helvetica", 26))
        self.username_entry.pack(pady=10)

        # Bind the Enter key to the username_entry widget
        self.username_entry.bind("<Return>", self.on_enter)

        # Create the numpad layou
        # t with larger font and padding
        button_frame = tk.Frame(self)
        button_frame.pack(side="top", pady=10)
        button_grid = [
            ['     1     ', '      2      ', '     3     '],
            ['     4     ', '      5      ', '     6     '],
            ['     7     ', '      8      ', '     9     '],
            ['     0      ', 'Clear', 'Login']
        ]

        for row_index, row in enumerate(button_grid):
            for col_index, number in enumerate(row):
                if number == 'Clear':
                    button = tk.Button(button_frame, text='🗑️ ' + number, bg='red', fg='white',
                                       command=self.handle_clear_button_click,
                                       font=("Helvetica", 20))
                elif number == 'Login':
                    button = tk.Button(button_frame, text='🔑 ' + number, bg='green', fg='white', font=("Helvetica", 20),
                                       command=lambda: controller.validate_login(self.username_entry.get()))
                else:
                    button = tk.Button(button_frame, text=number,
                                       command=lambda n=number: self.handle_num_button_click(n), bg='white',
                                       font=("Helvetica", 30))
                button.grid(row=row_index, column=col_index, sticky="nsew", padx=5,
                            pady=5)  # Increased padding between buttons

    def on_enter(self, event):
        self.controller.validate_login(self.username_entry.get())

    # Define functions for numpad interaction
    def handle_num_button_click(self, number):
        self.username_entry.insert(tk.END, number.strip())

    def handle_clear_button_click(self):
        self.username_entry.delete(0, tk.END)


class AdminPage(tk.Frame):  # for the admin to add books

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Add a new Book to Library", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)

        subtitle1_label = tk.Label(self, text="Please scan the book's barcode using the barcode scanner:",
                                   font=('Helvetica', 20))
        subtitle1_label.pack(side="top", fill="x", pady=5)

        self.barcode_entry = tk.Entry(self, width=20, font=("Helvetica", 26))
        self.barcode_entry.pack(pady=10)

        subtitle2_label = tk.Label(self, text="Make sure to fill out the Book Name in the database after scanning",
                                   font=('Helvetica', 20))
        subtitle2_label.pack(side="top", fill="x", pady=5)

        logout_button = tk.Button(self, text='🔙 ' + "Logout",
                                  command=lambda: controller.logout(),
                                  font=('Helvetica', 30))
        logout_button.pack(pady=10)

        # Bind the Enter key to the barcode_entry widget
        self.barcode_entry.bind("<Return>", self.on_enter)

    def on_enter(self, event):
        self.controller.add_book(self.barcode_entry.get())


class MainUserPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="🤓 Hello, user" + ' 📚', font=('Helvetica', 35, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)

        button_frame = tk.Frame(self)
        button_frame.pack(side="top", pady=10)
        borrow_book_button = tk.Button(button_frame,
                                       text="Borrow" + " (➕📖) ",
                                       command=lambda: controller.goto_borrow_book_page(user_id=self.user_id),
                                       bg="green", fg='white', font=('Helvetica', 20, 'bold'),
                                       width=15, height=8)
        borrow_book_button.pack(side="left", padx=8)

        # Return button
        return_book_button = tk.Button(button_frame,
                                       text="Return" + " (➖📗) ",
                                       command=lambda: controller.goto_return_book_page(),
                                       font=('Helvetica', 20, 'bold'),
                                       bg="red", fg='white', width=15, height=8)
        return_book_button.pack(side="left", padx=8)

        # History button
        view_transactions_button = tk.Button(button_frame,
                                             text="My Books" + " 📑",
                                             command=lambda: controller.goto_user_status_page(
                                                 user_id=self.user_id,
                                                 prev_page="MainUserPage"),
                                             bg="blue", fg='white',
                                             font=('Helvetica', 20, 'bold'),
                                             width=15, height=8)
        view_transactions_button.pack(side="left", padx=8)

        # Logout button
        logout_button = tk.Button(self, text="Logout" + " 👋",
                                  command=lambda: controller.logout(), bg='black', fg='white',
                                  font=('Helvetica', 25, 'bold'))
        logout_button.pack(side="top", pady=(10, 10))


class UserStatusPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Books You've Borrowed", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)

        list_frame = tk.Frame(self)
        list_frame.pack(side='top', fill="x")

        self.user_book_names_listbox = tk.Listbox(
            list_frame,
            exportselection=False,
            selectmode=tk.SINGLE,
            font=('Helvetica', 20, 'bold'))

        #self.user_book_names_listbox.pack(side=tk.LEFT, fill=tk.Y)
        self.user_book_names_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create the second listbox
        self.user_date_listbox = tk.Listbox(list_frame,
                                            exportselection=False,
                                            selectmode=tk.SINGLE,
                                            font=('Helvetica', 20, 'bold'))
        self.user_date_listbox.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Create a shared scrollbar
        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL,
                                 command=lambda *args: (self.user_book_names_listbox.yview(*args),
                                                        self.user_date_listbox.yview(*args)))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Connect listboxes to the scrollbar
        self.user_book_names_listbox.config(yscrollcommand=scrollbar.set)
        self.user_date_listbox.config(yscrollcommand=scrollbar.set)

        # Create up and down buttons
        buttom_frame = tk.Frame(self)
        buttom_frame.pack(side='top', fill="x")
        button_up = tk.Button(
            buttom_frame,
            text="⬆️",
            command=self.scroll_up,
            font=('Helvetica', 30)
        )
        button_down = tk.Button(
            buttom_frame,
            text="⬇️",
            command=self.scroll_down,
            font=('Helvetica', 30)
        )
        button_down.pack(side=tk.LEFT, padx=10)

        button_go_back = tk.Button(
        buttom_frame, text = '🔙 ' + "Go Back",
        command = lambda: controller.show_frame(self.back_page),
        font = ('Helvetica', 30))
        button_go_back.pack(side=tk.LEFT, padx=200)

        button_up.pack(side=tk.RIGHT, padx=10)

        # self.back_page = "userInfoPage"
        # back_button = tk.Button(self, text='🔙 ' + "Go Back",
        #                         command=lambda: controller.show_frame(self.back_page),
        #                         font=('Helvetica', 30))
        # back_button.pack(pady=10)

    def scroll_up(self):
        self.user_book_names_listbox.yview("scroll", "-1", "units")
        self.user_date_listbox.yview("scroll", "-1", "units")

    def scroll_down(self):
        self.user_book_names_listbox.yview("scroll", "1", "units")
        self.user_date_listbox.yview("scroll", "1", "units")


class LoadingPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Loading, Please Wait...", font=('Helvetica', 25, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class StartingUpPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Starting Up, Please Wait...", font=('Helvetica', 25, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class IDScanLoadingPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Logging in, Please Wait...", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class BorrowBookLoadingPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Registering Book Borrow, Please Wait...", font=('Helvetica', 25, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class ReturnBookLoadingPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Registering Book Return, Please Wait...", font=('Helvetica', 25, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class NotificationPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.title_label = tk.Label(self, text="")
        self.title_label.pack(side="top", fill="x", pady=10)


class TransactionsLoadingPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Fetching Data, Please Wait...", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)


class BorrowBookPage(tk.Frame):  # for the user

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Borrow A Book", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)

        subtitle_label = tk.Label(self, text="Please scan the book's barcode using the barcode scanner:",
                                  font=('Helvetica', 20, "bold"))
        subtitle_label.pack(side="top", fill="x", pady=5)

        self.barcode_entry = tk.Entry(self, width=20, font=("Helvetica", 26))
        self.barcode_entry.pack(pady=10)

        self.user_id = ""

        back_button = tk.Button(self, text='🔙 ' + "Go Back",
                                command=lambda: controller.show_frame("MainUserPage"),
                                font=('Helvetica', 30))
        back_button.pack(pady=10)

        # Bind the Enter key to the barcode_entry widget
        self.barcode_entry.bind("<Return>", self.on_enter)

    def on_enter(self, event):
        self.controller.borrow_book(self.barcode_entry.get(), self.user_id)


class ReturnBookPage(tk.Frame):  # for the user

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        title_label = tk.Label(self, text="Return A Book", font=('Helvetica', 30, 'bold'))
        title_label.pack(side="top", fill="x", pady=10)

        subtitle_label = tk.Label(self, text="Please scan the book's barcode using the barcode scanner:",
                                  font=('Helvetica', 20, 'bold'))
        subtitle_label.pack(side="top", fill="x", pady=10)

        self.barcode_entry = tk.Entry(self, width=20, font=("Helvetica", 26))
        self.barcode_entry.pack(pady=10)

        back_button = tk.Button(self, text='🔙 ' + "Go Back",
                                command=lambda: controller.show_frame("MainUserPage"),
                                font=('Helvetica', 30))
        back_button.pack(pady=10)

        # Bind the Enter key to the barcode_entry widget
        self.barcode_entry.bind("<Return>", self.on_enter)

    def on_enter(self, event):
        self.controller.return_book(self.barcode_entry.get())


class AutomaticLogOutPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        line1 = tk.Label(self, text="We Noticed You've Been Logged In For A While,",
                         font=('Helvetica', 25, 'bold'))

        line1.pack(side="top", fill="x", pady=10)

        line2 = tk.Label(self, text="Are You Still There? Please Confirm Or You'll",
                         font=('Helvetica', 17, 'bold'))

        line2.pack(side="top", fill="x", pady=10)

        line3 = tk.Label(self, text="Be Logged Out Automatically In 15 Seconds",
                         font=('Helvetica', 17, 'bold'))

        line3.pack(side="top", fill="x", pady=10)

        # Variable to store the page when this message pops up:
        self.prev_page = ""

        # Variable to store the scheduled logout job:
        self.sched_logout = None

        # Create and pack 'Yes' and 'No' buttons
        button_frame = tk.Frame(self)
        button_frame.pack(pady=30)

        yes_button = tk.Button(button_frame, text="Yes, still here!" + " ✅", command=self.stay,
                               fg="green", font=('Helvetica', 25, 'bold'))
        yes_button.pack(side="left", padx=10)

    def stay(self):
        # here cancel the scheduled logout
        self.after_cancel(id=self.sched_logout)
        self.controller.show_frame(self.prev_page)

    def leave(self):
        # here cancel the scheduled logout
        self.after_cancel(id=self.sched_logout)
        self.controller.logout()


class StatusBar(tk.Frame):
    def __init__(self, parent):
        tk.Frame.__init__(self, parent, height=50)  # Increased height for better visibility
        self.pack(side="top", fill="x")

        # Time Label
        self.time_label = tk.Label(self, font=("Helvetica", 20))  # Increased font size
        self.time_label.pack(side="left", padx=20, pady=(30, 0))  # Increased padding

        # WI-FI connection Label
        self.connection_status_label = tk.Label(self, font=("Helvetica", 20))  # Increased font size
        self.connection_status_label.pack(side="right", padx=20, pady=(35, 0))  # Increased padding

        # Update time and date every second
        self.update_time()

    def update_time(self):
        now = datetime.datetime.now()
        self.time_label.config(text=now.strftime("%H:%M"))
        # self.date_label.config(text=now.strftime("%Y-%m-%d"))
        self.after(60 * 1000, self.update_time)  # Update every minute


if __name__ == "__main__":
    app = SampleApp()
    app.mainloop()